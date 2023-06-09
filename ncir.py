import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from flask import Blueprint, render_template, request


ncir=Blueprint('ncir', __name__, template_folder='templates')

count=0
path='C:/Users/roranjan/Downloads' 
final_file=path+'/HC.xlsx' 
def process():
    # Get the uploaded files from the request
    log_file = request.files['file']
    srmtctl_log_file = request.files['file']
    
    # Save the files to temporary locations
    log_file_path = 'temp/' + log_file.filename
    srmtctl_log_file_path = 'temp/' + srmtctl_log_file.filename
    log_file.save(log_file_path)
    srmtctl_log_file.save(srmtctl_log_file_path)



    
#code for ip address-
    v=open(path+'/ip.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool lan print' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+word[0]+', ,')
                if('IP Address              :' in data):
                    word=data.split(': ')
                    ss=word[1].split('\n')
                    sq=ss[0]
                    v.write(sq)
                if('cmd.run' in data):
                    break
    v.close()
 
    textdatafile = pd.read_csv(path+'/ip.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','NADCM NAME','IP']
    textdatafile.to_excel(path+'/ip.xlsx', index = None)
    os.remove(path+'/ip.txt')


#code for board product
    v=open(path+'/hw.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool fru print"' in data):
            while data:
                data=f.readline()
        if('compute' in data or 'storage' in data or 'controller' in data):
            word=data.split(':')
            sp=word[0]
            v.write('\n'+word[0]+',')
        if('Board Product         :' in data):
            word=data.split(': ')
            ss=word[1].split('\n')
            sq=ss[0]
            v.write(sq)
        if('cmd.run' in data):
            break
    v.close()
 
    textdatafile = pd.read_csv(path+'/hw.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','HW']
    textdatafile.to_excel(path+'/hw.xlsx', index = None)
    os.remove(path+'/hw.txt')

#code for BiOS version
    v=open(path+'/bios.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool mc getsysinfo system_fw_version' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+word[0]+',')
                if('BIOS Version:' in data):
                    word=data.split(':')
                    ss=word[1].split(' ')
                    sq=ss[0]
                    v.write(sq)
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/bios.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','BIOS Version']
    textdatafile.to_excel(path+'/bios.xlsx', index = None)
    os.remove(path+'/bios.txt')


#code firmware revision
    v=open(path+'/fw.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool mc info' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+word[0]+',')
                if('Firmware Revision         :' in data):
                    word=data.split(':')
                    ss=word[1].split('\n')
                    sq=ss[0]
                    v.write(sq)
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/fw.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','FW']
    textdatafile.to_excel(path+'/fw.xlsx', index = None)
    os.remove(path+'/fw.txt')


#code for sensor-
    v=open(path+'/sensor.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool sensor"' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+sp+',')
                    count+=1
                    flag=0
                    flag2=0
                    for i in range(22):
                        data=f.readline()
                        if('Get Device ID command failed' in data):
                            flag2=1
                            break
                        value=data.split('|')
                        if('ok' in value[3]):
                            flag=0
                        else:
                            flag=1
                    if(flag==0 and flag2==0):
                        v.write('ok')
                    else:  
                        v.write('check')
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/sensor.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','sensors']
    textdatafile.to_excel(path+'/sensor.xlsx', index = None)
    os.remove(path+'/sensor.txt')



#code for ce_count-
    v=open(path+'/ce_count.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "grep "[0-9]" /sys/devices/system/edac/mc/mc*/ce_count' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+sp+',')
                    flag=0
                    for i in range(4):
                        data=f.readline()
                        value=data.split(':')
                        p=value[1]
                        s=int(p)
                        if(s>0):
                            flag=1
                    if(flag==0):
                        v.write('ok')
                    else:
                        v.write('check')
                    
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/ce_count.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','ce_count']
    textdatafile.to_excel(path+'/ce_count.xlsx', index = None)
    os.remove(path+'/ce_count.txt')



#code for ue_count-
    v=open(path+'/ue_count.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cat /sys/devices/system/edac/mc/mc*/ue_count' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+sp+',')
                    flag=0
                    for i in range(4):
                        data=f.readline()
                        s=int(data)
                        if(s>0):
                            flag=1
                    if(flag==0):
                        v.write('ok')
                    else:
                        v.write('check')
                    
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/ue_count.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','ue_count']
    textdatafile.to_excel(path+'/ue_count.xlsx', index = None)
    os.remove(path+'/ue_count.txt')



#code for fans-
    v=open(path+'/fan.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool chassis status' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+word[0]+',')
                if('Cooling/Fan Fault    :' in data):
                    word=data.split(':')
                    if('false' in word[1]):
                        sq="ok"
                    else:
                        sq="nok"
                    v.write(sq)
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/fan.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','Fans']
    textdatafile.to_excel(path+'/fan.xlsx', index = None)
    os.remove(path+'/fan.txt')


#code PSUs
    v=open(path+'/psu.txt','a+')
    with open(log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('cmd.run "ipmitool sdr"' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'storage' in data or 'controller' in data):
                    word=data.split(':')
                    sp=word[0]
                    v.write('\n'+word[0]+',')
                if('Power            |' in data):
                    word=data.split('|')
                    if('ok' in word[2]):
                        v.write("ok")
                    else:
                        v.write("nok")
                if('cmd.run' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/psu.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','PSUs']
    textdatafile.to_excel(path+'/psu.xlsx', index = None)
    os.remove(path+'/psu.txt')


#code for manufacturer_sda-
    v=open(path+'/manu.txt','a+')
    with open(srmtctl_log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('sudo smartctl -a /dev/sda' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'controller' in data):
                    word=data.split(':')
                    v.write(word[0]+',')
                if('SMART overall-health self-assessment test result:' in data):
                    w=data.split(':')
                    ss=w[1]
                    if('PASSED' in ss):
                        sw='ok'
                    else:
                        sw='nok'
                    v.write(sw+',')
                if('Model Family:' in data):
                    wor=data.split(':     ')
                    sc=wor[1].split('\n')
                    sq=sc[0]
                    v.write(sq+',')
                if('Reallocated_Sector_Ct' in data):
                    wo=data.split('       ')
                    sp=wo[2].split('\n')
                    ss=sp[0]
                    if('0' in ss):
                        sd='ok'
                    else:
                        sd=ss+' in disk'
                    v.write(sd+'\n')
                if('Smartctl open device: /dev/sdb failed: No such device' in data):
                    v.write('\n')
                if('sudo smartctl -a /dev/sd' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/manu.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','Manufacturer sda disk','disk sda(overall)','disk sda(reallocated_sector_ct)']
    textdatafile.to_excel('C:/Users/roranjan/Downloads/manu.xlsx', index = None)
    os.remove(path+'/manu.txt')


#code for manufacturer_sdb-
    v=open(path+'/manu2.txt','a+')
    with open(srmtctl_log_file,'r') as f:
        data=True
    while data:
        data=f.readline()
        if('sudo smartctl -a /dev/sdb' in data):
            while data:
                data=f.readline()
                if('compute' in data or 'controller' in data):
                    word=data.split(':')
                    v.write(word[0]+',')
                if('SMART overall-health self-assessment test result:' in data):
                    w=data.split(':')
                    ss=w[1]
                    if('PASSED' in ss):
                        sw='ok'
                    else:
                        sw='nok'
                    v.write(sw+',')
                if('Model Family:' in data):
                    wor=data.split(':     ')
                    sc=wor[1].split('\n')
                    sq=sc[0]
                    v.write(sq+',')
                if('Reallocated_Sector_Ct' in data):
                    wo=data.split('       ')
                    sp=wo[2].split('\n')
                    ss=sp[0]
                    if('0' in ss):
                        sd='ok'
                    else:
                        sd=ss+' in disk'
                    v.write(sd+'\n')
                if('Smartctl open device: /dev/sdb failed: No such device' in data):
                    v.write('\n')
                if('sudo smartctl -a /dev/sd' in data or 'stack@' in data):
                    break
    v.close()

    textdatafile = pd.read_csv(path+'/manu2.txt',header=None,on_bad_lines='skip')
    textdatafile.columns=['CBIS NAME','Manufacturer sdb disk','disk sdb(overall)','disk sdb(reallocated_sector_ct)']
    textdatafile.to_excel('C:/Users/roranjan/Downloads/manu2.xlsx', index = None)
    os.remove(path+'/manu2.txt')


    #code to create excel file-
df1 = pd.read_excel(path+'/ip.xlsx')
df2 = pd.read_excel(path+'/hw.xlsx')
df3 = pd.read_excel(path+'/bios.xlsx')
df4 = pd.read_excel(path+'/fw.xlsx')
df5 = pd.read_excel(path+'/fan.xlsx')
df6 = pd.read_excel(path+'/psu.xlsx')
df7 = pd.read_excel(path+'/manu.xlsx')
df8 = pd.read_excel(path+'/manu2.xlsx')
df9 = pd.read_excel(path+'/ue_count.xlsx')
df10 = pd.read_excel(path+'/ce_count.xlsx')
df11 = pd.read_excel(path+'/sensor.xlsx')
df_combine=df1.merge(df2, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df3, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df4, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df11, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df10, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df9, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df5, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df6, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df7, on='CBIS NAME', how='outer')
df_combine=df_combine.merge(df8, on='CBIS NAME', how='outer')
df_combine.to_excel(final_file, index = None)
os.remove(path+'/ip.xlsx')
os.remove(path+'/hw.xlsx')
os.remove(path+'/bios.xlsx')
os.remove(path+'/fw.xlsx')
os.remove(path+'/fan.xlsx')
os.remove(path+'/psu.xlsx')
os.remove(path+'/manu.xlsx')
os.remove(path+'/manu2.xlsx')
os.remove(path+'/ue_count.xlsx')
os.remove(path+'/ce_count.xlsx')
os.remove(path+'/sensor.xlsx')


#code for sorting the excel_sheet-
df=pd.read_excel(final_file)
df_sor = df.sort_values(by="CBIS NAME")
df_sor.to_excel(final_file,index=None)

#code to apply color in excel
df=pd.read_excel(final_file)
def color_rule1(tag):
    var1=df['Fans']
    return['background-color: red' if x=='nok' else 'background-color: #90EE90' for x in var1]
def color_rule2(tag):
    var1=df['PSUs']
    return['background-color: red' if x=='nok' else 'background-color: #90EE90' for x in var1]
def color_rule3(tag):
    var1=df['CBIS NAME']
    return['border-style:solid' if 'overcloud' in x else 'border-style: solid' for x in var1]
def color_rule4(tag):
    var1=df['disk sda(overall)']
    return['background-color: red' if x=='nok' else 'background-color: #90EE90' for x in var1]
def color_rule5(tag):
    var1=df['disk sdb(overall)']
    return['background-color: red' if x=='nok' else 'background-color: #90EE90' for x in var1]
def color_rule6(tag):
    var1=df['disk sda(reallocated_sector_ct)']
    return['background-color: #90EE90' if x=='ok' else 'background-color: red' for x in var1]
def color_rule7(tag):
    var1=df['disk sdb(reallocated_sector_ct)']
    return['background-color: #90EE90' if x=='ok' else 'background-color: red' for x in var1]
def color_rule8(tag):
    var1=df['ue_count']
    return['background-color: red' if x=='check' else 'background-color: #90EE90' for x in var1]
def color_rule9(tag):
    var1=df['ce_count']
    return['background-color: red' if x=='check' else 'background-color: #90EE90' for x in var1]
def color_rule10(tag):
    var1=df['sensors']
    return['background-color: red' if x=='check' else 'background-color: #90EE90' for x in var1]
df.style.apply(color_rule10, subset=['sensors']).apply(color_rule1, subset=['Fans']).apply(color_rule2, subset=['PSUs']).apply(color_rule3).apply(color_rule4, subset=['disk sda(overall)']).apply(color_rule5, subset=['disk sdb(overall)']).apply(color_rule6, subset=['disk sda(reallocated_sector_ct)']).apply(color_rule7, subset=['disk sdb(reallocated_sector_ct)']).apply(color_rule8, subset=['ue_count']).apply(color_rule9, subset=['ce_count']).to_excel(final_file, index=None) 
wb=openpyxl.load_workbook(final_file)
ws=wb['Sheet1']
fill_pattern=PatternFill(patternType='solid',fgColor='ADD8E6')
ws['A1'].fill=fill_pattern
ws['A1'].alignment = Alignment(horizontal='left')
ws['B1'].fill=fill_pattern
ws['B1'].alignment = Alignment(horizontal='left')
ws['C1'].fill=fill_pattern
ws['C1'].alignment = Alignment(horizontal='left')
ws['D1'].fill=fill_pattern
ws['D1'].alignment = Alignment(horizontal='left')
ws['E1'].fill=fill_pattern
ws['E1'].alignment = Alignment(horizontal='left')
ws['F1'].fill=fill_pattern
ws['F1'].alignment = Alignment(horizontal='left')
ws['G1'].fill=fill_pattern
ws['G1'].alignment = Alignment(horizontal='left')
ws['H1'].fill=fill_pattern
ws['H1'].alignment = Alignment(horizontal='left')
ws['I1'].fill=fill_pattern
ws['I1'].alignment = Alignment(horizontal='left')
ws['J1'].fill=fill_pattern
ws['J1'].alignment = Alignment(horizontal='left')
ws['K1'].fill=fill_pattern
ws['K1'].alignment = Alignment(horizontal='left')
ws['L1'].fill=fill_pattern
ws['L1'].alignment = Alignment(horizontal='left')
ws['M1'].fill=fill_pattern
ws['M1'].alignment = Alignment(horizontal='left')
ws['N1'].fill=fill_pattern
ws['N1'].alignment = Alignment(horizontal='left')
ws['O1'].fill=fill_pattern
ws['O1'].alignment = Alignment(horizontal='left')
ws['P1'].fill=fill_pattern
ws['P1'].alignment = Alignment(horizontal='left')
ws['Q1'].fill=fill_pattern
ws['Q1'].alignment = Alignment(horizontal='left')
wb.save(final_file)
print('File created successfully...')